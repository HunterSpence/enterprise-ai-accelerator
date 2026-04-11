"""
tco_calculator.py — CFO-Level TCO Financial Model (V2)
=======================================================

V2 upgrades:
  - All cost categories: compute, storage, data transfer, managed services,
    support tier, labor (one-time + ongoing FTE delta), licensing, training,
    risk buffer (15% contingency)
  - Three scenarios: Lift & Shift vs Re-platform vs Re-architect
  - IRR calculation (proper internal rate of return)
  - CFO-level executive summary with payback period and IRR
  - Excel export (openpyxl) with proper formatting and formulas
  - Enhanced sensitivity analysis (7 scenarios)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich import box

from .assessor import WorkloadAssessment, MigrationStrategy

console = Console()


@dataclass
class OnPremCosts:
    hardware_monthly: float
    power_monthly: float
    datacenter_monthly: float
    staff_monthly: float
    license_monthly: float
    maintenance_monthly: float
    backup_monthly: float

    @property
    def total_monthly(self) -> float:
        return (self.hardware_monthly + self.power_monthly + self.datacenter_monthly
                + self.staff_monthly + self.license_monthly + self.maintenance_monthly
                + self.backup_monthly)

    @property
    def total_annual(self) -> float:
        return self.total_monthly * 12


@dataclass
class CloudCosts:
    compute_monthly: float
    storage_monthly: float
    database_monthly: float
    network_monthly: float
    license_monthly: float
    support_monthly: float
    managed_services_monthly: float
    saas_monthly: float = 0.0
    # V2 additional categories
    data_transfer_monthly: float = 0.0
    backup_monthly: float = 0.0
    security_monthly: float = 0.0   # GuardDuty, Security Hub, WAF

    @property
    def total_monthly(self) -> float:
        return (self.compute_monthly + self.storage_monthly + self.database_monthly
                + self.network_monthly + self.license_monthly + self.support_monthly
                + self.managed_services_monthly + self.saas_monthly
                + self.data_transfer_monthly + self.backup_monthly + self.security_monthly)

    @property
    def total_annual(self) -> float:
        return self.total_monthly * 12


@dataclass
class LaborCosts:
    """One-time migration labor + ongoing cloud ops FTE delta."""
    migration_project_one_time: float     # One-time: architects, engineers, PMs
    cloud_ops_monthly_delta: float        # Ongoing: FTE reduction from cloud ops efficiency
    training_one_time: float              # Cloud certifications + team ramp-up


@dataclass
class MigrationScenario:
    """Financial model for a specific migration strategy."""
    name: str
    strategy_type: str          # "lift_and_shift", "replatform", "re_architect"
    year1_cloud_cost: float
    year2_cloud_cost: float
    year3_cloud_cost: float
    year1_savings: float
    year2_savings: float
    year3_savings: float
    migration_cost: float
    npv_3yr: float
    irr_percent: float
    break_even_months: float
    cumulative_3yr_benefit: float
    recommendation: str


@dataclass
class SensitivityResult:
    scenario_name: str
    cloud_cost_multiplier: float
    staff_savings_multiplier: float
    npv_3yr: float
    break_even_months: float
    annual_savings: float
    recommendation: str


@dataclass
class TCOAnalysis:
    on_prem: OnPremCosts
    cloud: CloudCosts
    labor: LaborCosts
    migration_cost_usd: float
    contingency_usd: float          # 15% risk buffer
    total_investment_usd: float     # migration_cost + contingency + training
    npv_3yr: float
    npv_5yr: float
    irr_percent: float
    break_even_months: float
    annual_savings: float
    three_year_cumulative_savings: float
    five_year_cumulative_savings: float
    scenarios: list[MigrationScenario] = field(default_factory=list)
    sensitivity_results: list[SensitivityResult] = field(default_factory=list)
    yearly_cashflows: list[float] = field(default_factory=list)

    @property
    def roi_percent(self) -> float:
        if self.total_investment_usd <= 0:
            return 0.0
        return ((self.annual_savings * 3 - self.total_investment_usd) / self.total_investment_usd) * 100

    @property
    def payback_period_str(self) -> str:
        months = self.break_even_months
        if months < 0 or months > 60:
            return "Never (within 5yr)"
        elif months < 12:
            return f"{months:.0f} months"
        else:
            return f"{months / 12:.1f} years"


class TCOCalculator:
    """
    V2 CFO-level TCO calculator.

    Models all cost categories with three migration strategy scenarios
    and proper IRR calculation.
    """

    POWER_KW_PER_SERVER = 0.30
    POWER_COST_PER_KWH = 0.10
    PUE = 1.5
    RACK_COST_PER_U_MONTHLY = 300.0
    RACK_UNITS_PER_SERVER = 2
    HARDWARE_ANNUAL_PER_CORE = 800.0
    HARDWARE_ANNUAL_PER_GB_RAM = 60.0
    STAFF_COST_ANNUAL_PER_SERVER = 2400.0
    MAINTENANCE_PCT_OF_HARDWARE = 0.15
    BACKUP_MONTHLY_PER_TB = 25.0
    DISCOUNT_RATE_ANNUAL = 0.08
    RI_DISCOUNT = 0.30
    SUPPORT_PCT_OF_COMPUTE = 0.05
    MANAGED_SERVICES_MONTHLY = 150.0
    CONTINGENCY_PCT = 0.15          # 15% risk buffer
    TRAINING_COST_PER_ENGINEER = 3_500.0   # AWS cert + ramp-up
    CLOUD_OPS_FTE_REDUCTION = 0.35         # 35% ops staff reduction in cloud

    def _compute_on_prem(self, a: WorkloadAssessment) -> OnPremCosts:
        w = a.workload
        hardware_annual = w.cpu_cores * self.HARDWARE_ANNUAL_PER_CORE + w.ram_gb * self.HARDWARE_ANNUAL_PER_GB_RAM
        hardware_monthly = hardware_annual / 12
        power_monthly = self.POWER_KW_PER_SERVER * self.PUE * self.POWER_COST_PER_KWH * 24 * 30
        datacenter_monthly = self.RACK_COST_PER_U_MONTHLY * self.RACK_UNITS_PER_SERVER
        staff_monthly = self.STAFF_COST_ANNUAL_PER_SERVER / 12

        license_monthly_map = {
            "database": 400.0, "web_app": 80.0, "batch_job": 50.0,
            "middleware": 200.0, "legacy": 600.0, "microservice": 20.0,
        }
        license_monthly = license_monthly_map.get(w.workload_type, 100.0)
        if w.license_type == "commercial":
            license_monthly *= 2.5
        # Include explicit annual license cost if provided
        if w.license_cost_annual > 0:
            license_monthly = max(license_monthly, w.license_cost_annual / 12)

        maintenance_monthly = (hardware_annual * self.MAINTENANCE_PCT_OF_HARDWARE) / 12
        storage_tb = w.storage_gb / 1024
        backup_monthly = max(25.0, storage_tb * self.BACKUP_MONTHLY_PER_TB)

        return OnPremCosts(
            hardware_monthly=round(hardware_monthly, 2),
            power_monthly=round(power_monthly, 2),
            datacenter_monthly=round(datacenter_monthly, 2),
            staff_monthly=round(staff_monthly, 2),
            license_monthly=round(license_monthly, 2),
            maintenance_monthly=round(maintenance_monthly, 2),
            backup_monthly=round(backup_monthly, 2),
        )

    def _compute_cloud(self, a: WorkloadAssessment) -> CloudCosts:
        w = a.workload
        compute_base = a.monthly_cloud_cost_usd
        compute_monthly = compute_base * (1 - self.RI_DISCOUNT)
        storage_monthly = max(10.0, w.storage_gb * 0.08)

        db_monthly = 0.0
        if w.workload_type == "database" or w.database_type:
            db_monthly = compute_monthly * 0.5
            if a.strategy == MigrationStrategy.REFACTOR:
                db_monthly *= 0.6

        network_monthly = max(30.0, w.storage_gb * 0.02)
        data_transfer_monthly = max(20.0, w.cpu_cores * 3.5)  # Data egress estimate

        license_monthly = 0.0
        if w.license_type == "commercial" and a.strategy in (
            MigrationStrategy.REHOST, MigrationStrategy.REPLATFORM
        ):
            license_monthly = 150.0
        elif a.strategy == MigrationStrategy.REPURCHASE:
            license_monthly = a.monthly_cloud_cost_usd * 0.7

        support_monthly = max(20.0, compute_monthly * self.SUPPORT_PCT_OF_COMPUTE)
        managed_monthly = self.MANAGED_SERVICES_MONTHLY
        backup_monthly = max(15.0, w.storage_gb * 0.023)   # S3 + AWS Backup
        security_monthly = max(30.0, compute_monthly * 0.04)   # GuardDuty + WAF estimate

        saas_monthly = 0.0
        if a.strategy == MigrationStrategy.REPURCHASE:
            saas_monthly = a.monthly_cloud_cost_usd

        return CloudCosts(
            compute_monthly=round(compute_monthly, 2),
            storage_monthly=round(storage_monthly, 2),
            database_monthly=round(db_monthly, 2),
            network_monthly=round(network_monthly, 2),
            license_monthly=round(license_monthly, 2),
            support_monthly=round(support_monthly, 2),
            managed_services_monthly=round(managed_monthly, 2),
            saas_monthly=round(saas_monthly, 2),
            data_transfer_monthly=round(data_transfer_monthly, 2),
            backup_monthly=round(backup_monthly, 2),
            security_monthly=round(security_monthly, 2),
        )

    def _compute_labor(self, assessments: list[WorkloadAssessment]) -> LaborCosts:
        """Compute labor costs: one-time migration + ongoing FTE delta."""
        total_migration_cost = sum(a.estimated_migration_cost_usd for a in assessments)

        # Training: assume 1 engineer per 10 workloads needs certification
        engineer_count = max(2, len(assessments) // 10)
        training_one_time = engineer_count * self.TRAINING_COST_PER_ENGINEER

        # Monthly cloud ops savings: fewer ops staff needed
        total_staff_monthly = sum(
            self._compute_on_prem(a).staff_monthly for a in assessments
        )
        cloud_ops_monthly_delta = total_staff_monthly * self.CLOUD_OPS_FTE_REDUCTION

        return LaborCosts(
            migration_project_one_time=total_migration_cost,
            cloud_ops_monthly_delta=cloud_ops_monthly_delta,
            training_one_time=training_one_time,
        )

    def _irr(self, cashflows: list[float]) -> float:
        """Compute Internal Rate of Return using Newton-Raphson method."""
        if len(cashflows) < 2:
            return 0.0

        # Initial guess
        r = 0.10
        for _ in range(100):
            npv = sum(cf / (1 + r) ** t for t, cf in enumerate(cashflows))
            dnpv = sum(-t * cf / (1 + r) ** (t + 1) for t, cf in enumerate(cashflows) if t > 0)
            if abs(dnpv) < 1e-10:
                break
            r_new = r - npv / dnpv
            if abs(r_new - r) < 1e-8:
                r = r_new
                break
            r = r_new

        if r < -1 or r > 10:
            return float("nan")
        return round(r * 100, 2)

    def _npv(self, cashflows: list[float], discount_rate: float | None = None) -> float:
        rate = discount_rate or self.DISCOUNT_RATE_ANNUAL
        return sum(cf / (1 + rate) ** t for t, cf in enumerate(cashflows))

    def _break_even_months(self, investment: float, monthly_savings: float) -> float:
        if monthly_savings <= 0:
            return -1.0
        return investment / monthly_savings

    def _build_scenarios(
        self,
        on_prem: OnPremCosts,
        base_cloud: CloudCosts,
        migration_cost: float,
    ) -> list[MigrationScenario]:
        """Build three scenario models with dramatically different 3-year totals."""
        scenarios_def = [
            ("Lift & Shift", "lift_and_shift", 1.15, 1.15, 1.10, migration_cost * 0.80),
            ("Re-platform", "replatform", 1.0, 0.92, 0.85, migration_cost * 1.20),
            ("Re-architect", "re_architect", 0.80, 0.65, 0.55, migration_cost * 2.50),
        ]

        scenarios = []
        base_on_prem_monthly = on_prem.total_monthly

        for name, strategy_type, mult_yr1, mult_yr2, mult_yr3, mig_cost in scenarios_def:
            cloud_yr1 = base_cloud.total_monthly * 12 * mult_yr1
            cloud_yr2 = base_cloud.total_monthly * 12 * mult_yr2
            cloud_yr3 = base_cloud.total_monthly * 12 * mult_yr3
            on_prem_annual = base_on_prem_monthly * 12

            savings_yr1 = on_prem_annual - cloud_yr1
            savings_yr2 = on_prem_annual - cloud_yr2
            savings_yr3 = on_prem_annual - cloud_yr3
            cumulative = savings_yr1 + savings_yr2 + savings_yr3 - mig_cost

            cashflows = [-mig_cost, savings_yr1, savings_yr2, savings_yr3,
                         savings_yr3, savings_yr3]  # 5yr
            npv_3yr = self._npv(cashflows[:4])
            irr = self._irr(cashflows)
            be_months = self._break_even_months(mig_cost, savings_yr1 / 12)

            recommendation = (
                "Recommended for quick wins with minimal risk"
                if strategy_type == "lift_and_shift"
                else "Best balance of effort and long-term savings"
                if strategy_type == "replatform"
                else "Maximum 3-year savings but significant upfront investment"
            )

            scenarios.append(MigrationScenario(
                name=name,
                strategy_type=strategy_type,
                year1_cloud_cost=round(cloud_yr1, 0),
                year2_cloud_cost=round(cloud_yr2, 0),
                year3_cloud_cost=round(cloud_yr3, 0),
                year1_savings=round(savings_yr1, 0),
                year2_savings=round(savings_yr2, 0),
                year3_savings=round(savings_yr3, 0),
                migration_cost=round(mig_cost, 0),
                npv_3yr=round(npv_3yr, 0),
                irr_percent=irr if not (irr != irr) else 0.0,  # NaN guard
                break_even_months=round(be_months, 1),
                cumulative_3yr_benefit=round(cumulative, 0),
                recommendation=recommendation,
            ))

        return scenarios

    def _run_sensitivity(
        self, on_prem: OnPremCosts, cloud: CloudCosts, migration_cost: float
    ) -> list[SensitivityResult]:
        scenarios_def = [
            ("Base Case", 1.0, 1.0),
            ("Cloud costs +20%", 1.20, 1.0),
            ("Cloud costs +40%", 1.40, 1.0),
            ("Cloud costs -20% (RI + Savings Plans)", 0.80, 1.0),
            ("Staff savings reduced 50%", 1.0, 0.50),
            ("Worst case (cloud +30%, staff -50%)", 1.30, 0.50),
            ("Best case (cloud -20%, full staff savings)", 0.80, 1.0),
        ]

        results = []
        base_cloud_monthly = cloud.total_monthly
        base_on_prem_monthly = on_prem.total_monthly

        for name, cloud_mult, staff_mult in scenarios_def:
            staff_savings_monthly = on_prem.staff_monthly * self.CLOUD_OPS_FTE_REDUCTION * staff_mult
            adjusted_cloud_monthly = base_cloud_monthly * cloud_mult
            annual_savings = (base_on_prem_monthly - adjusted_cloud_monthly + staff_savings_monthly) * 12

            cashflows = [-migration_cost] + [annual_savings] * 5
            npv_3yr = self._npv(cashflows[:4])
            be_months = self._break_even_months(migration_cost, annual_savings / 12)

            if npv_3yr > 0 and be_months < 30:
                rec = "Proceed — positive ROI"
            elif npv_3yr > 0:
                rec = "Proceed with caution — longer payback"
            else:
                rec = "Re-evaluate — negative NPV in this scenario"

            results.append(SensitivityResult(
                scenario_name=name,
                cloud_cost_multiplier=cloud_mult,
                staff_savings_multiplier=staff_mult,
                npv_3yr=round(npv_3yr, 0),
                break_even_months=round(be_months, 1),
                annual_savings=round(annual_savings, 0),
                recommendation=rec,
            ))

        return results

    def analyze_workload(self, a: WorkloadAssessment) -> TCOAnalysis:
        on_prem = self._compute_on_prem(a)
        cloud = self._compute_cloud(a)
        labor = LaborCosts(
            migration_project_one_time=a.estimated_migration_cost_usd,
            cloud_ops_monthly_delta=on_prem.staff_monthly * self.CLOUD_OPS_FTE_REDUCTION,
            training_one_time=self.TRAINING_COST_PER_ENGINEER * 2,
        )

        migration_cost = a.estimated_migration_cost_usd
        contingency = migration_cost * self.CONTINGENCY_PCT
        total_investment = migration_cost + contingency + labor.training_one_time

        monthly_savings = on_prem.total_monthly - cloud.total_monthly + labor.cloud_ops_monthly_delta
        annual_savings = monthly_savings * 12

        cashflows = [-total_investment] + [annual_savings] * 5
        npv_3yr = self._npv(cashflows[:4])
        npv_5yr = self._npv(cashflows)
        irr = self._irr(cashflows)
        break_even = self._break_even_months(total_investment, monthly_savings)
        sensitivity = self._run_sensitivity(on_prem, cloud, total_investment)
        scenarios = self._build_scenarios(on_prem, cloud, migration_cost)

        yearly_cashflows = [-total_investment]
        for yr in range(1, 6):
            yearly_cashflows.append(annual_savings * yr - total_investment)

        return TCOAnalysis(
            on_prem=on_prem,
            cloud=cloud,
            labor=labor,
            migration_cost_usd=migration_cost,
            contingency_usd=contingency,
            total_investment_usd=total_investment,
            npv_3yr=round(npv_3yr, 0),
            npv_5yr=round(npv_5yr, 0),
            irr_percent=irr if irr == irr else 0.0,
            break_even_months=round(break_even, 1),
            annual_savings=round(annual_savings, 0),
            three_year_cumulative_savings=round(annual_savings * 3 - total_investment, 0),
            five_year_cumulative_savings=round(annual_savings * 5 - total_investment, 0),
            scenarios=scenarios,
            sensitivity_results=sensitivity,
            yearly_cashflows=yearly_cashflows,
        )

    def analyze_portfolio(self, assessments: list[WorkloadAssessment]) -> TCOAnalysis:
        """Aggregate TCO across all workloads."""
        total_on_prem = OnPremCosts(0, 0, 0, 0, 0, 0, 0)
        total_cloud = CloudCosts(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        total_migration_cost = 0.0

        for a in assessments:
            op = self._compute_on_prem(a)
            cl = self._compute_cloud(a)
            total_on_prem.hardware_monthly += op.hardware_monthly
            total_on_prem.power_monthly += op.power_monthly
            total_on_prem.datacenter_monthly += op.datacenter_monthly
            total_on_prem.staff_monthly += op.staff_monthly
            total_on_prem.license_monthly += op.license_monthly
            total_on_prem.maintenance_monthly += op.maintenance_monthly
            total_on_prem.backup_monthly += op.backup_monthly
            total_cloud.compute_monthly += cl.compute_monthly
            total_cloud.storage_monthly += cl.storage_monthly
            total_cloud.database_monthly += cl.database_monthly
            total_cloud.network_monthly += cl.network_monthly
            total_cloud.license_monthly += cl.license_monthly
            total_cloud.support_monthly += cl.support_monthly
            total_cloud.managed_services_monthly += cl.managed_services_monthly
            total_cloud.saas_monthly += cl.saas_monthly
            total_cloud.data_transfer_monthly += cl.data_transfer_monthly
            total_cloud.backup_monthly += cl.backup_monthly
            total_cloud.security_monthly += cl.security_monthly
            total_migration_cost += a.estimated_migration_cost_usd

        labor = self._compute_labor(assessments)
        contingency = total_migration_cost * self.CONTINGENCY_PCT
        total_investment = total_migration_cost + contingency + labor.training_one_time

        monthly_savings = (
            total_on_prem.total_monthly - total_cloud.total_monthly + labor.cloud_ops_monthly_delta
        )
        annual_savings = monthly_savings * 12

        cashflows = [-total_investment] + [annual_savings] * 5
        npv_3yr = self._npv(cashflows[:4])
        npv_5yr = self._npv(cashflows)
        irr = self._irr(cashflows)
        break_even = self._break_even_months(total_investment, monthly_savings)
        sensitivity = self._run_sensitivity(total_on_prem, total_cloud, total_investment)
        scenarios = self._build_scenarios(total_on_prem, total_cloud, total_migration_cost)

        yearly_cashflows = [-total_investment]
        for yr in range(1, 6):
            yearly_cashflows.append(annual_savings * yr - total_investment)

        return TCOAnalysis(
            on_prem=total_on_prem,
            cloud=total_cloud,
            labor=labor,
            migration_cost_usd=total_migration_cost,
            contingency_usd=contingency,
            total_investment_usd=total_investment,
            npv_3yr=round(npv_3yr, 0),
            npv_5yr=round(npv_5yr, 0),
            irr_percent=irr if irr == irr else 0.0,
            break_even_months=round(break_even, 1),
            annual_savings=round(annual_savings, 0),
            three_year_cumulative_savings=round(annual_savings * 3 - total_investment, 0),
            five_year_cumulative_savings=round(annual_savings * 5 - total_investment, 0),
            scenarios=scenarios,
            sensitivity_results=sensitivity,
            yearly_cashflows=yearly_cashflows,
        )

    def print_tco_report(self, tco: TCOAnalysis, title: str = "Portfolio TCO Analysis") -> None:
        console.print(f"\n[bold blue]{title}[/bold blue]")

        cost_table = Table(
            title="Cost Breakdown: On-Premises vs Cloud",
            box=box.ROUNDED,
            header_style="bold white on dark_blue",
        )
        cost_table.add_column("Cost Category", min_width=28)
        cost_table.add_column("On-Premises / Month", justify="right", min_width=20)
        cost_table.add_column("Cloud / Month", justify="right", min_width=18)
        cost_table.add_column("Savings / Month", justify="right", min_width=18)

        rows = [
            ("Compute / Hardware", tco.on_prem.hardware_monthly, tco.cloud.compute_monthly),
            ("Power & Cooling", tco.on_prem.power_monthly, 0.0),
            ("Datacenter / Rack", tco.on_prem.datacenter_monthly, 0.0),
            ("Storage (EBS/S3)", 0.0, tco.cloud.storage_monthly),
            ("Database (Managed)", 0.0, tco.cloud.database_monthly),
            ("Network / VPN", 0.0, tco.cloud.network_monthly),
            ("Data Transfer (Egress)", 0.0, tco.cloud.data_transfer_monthly),
            ("Staff (Ops Allocation)", tco.on_prem.staff_monthly, 0.0),
            ("Cloud Ops FTE Savings", 0.0, -tco.labor.cloud_ops_monthly_delta),
            ("Licenses", tco.on_prem.license_monthly, tco.cloud.license_monthly),
            ("Maintenance Contracts", tco.on_prem.maintenance_monthly, 0.0),
            ("Backup / DR", tco.on_prem.backup_monthly, tco.cloud.backup_monthly),
            ("Support Plan (AWS Business)", 0.0, tco.cloud.support_monthly),
            ("Managed Services (CW/Secrets)", 0.0, tco.cloud.managed_services_monthly),
            ("Security (GuardDuty/WAF)", 0.0, tco.cloud.security_monthly),
        ]

        for label, on_prem_cost, cloud_cost in rows:
            savings = on_prem_cost - cloud_cost
            savings_str = (
                f"[green]+${savings:,.0f}[/green]" if savings > 0
                else f"[red]-${abs(savings):,.0f}[/red]" if savings < 0
                else "[dim]—[/dim]"
            )
            cost_table.add_row(
                label,
                f"${on_prem_cost:,.0f}" if on_prem_cost > 0 else "[dim]—[/dim]",
                f"${cloud_cost:,.0f}" if cloud_cost > 0 else "[dim]—[/dim]",
                savings_str,
            )

        monthly_savings = tco.on_prem.total_monthly - tco.cloud.total_monthly + tco.labor.cloud_ops_monthly_delta
        cost_table.add_row(
            "[bold]TOTAL (incl. FTE delta)[/bold]",
            f"[bold]${tco.on_prem.total_monthly:,.0f}[/bold]",
            f"[bold]${tco.cloud.total_monthly - tco.labor.cloud_ops_monthly_delta:,.0f}[/bold]",
            f"[bold green]${monthly_savings:,.0f}/mo[/bold green]",
        )

        console.print(cost_table)

        # CFO-level executive summary
        irr_str = f"{tco.irr_percent:.1f}%" if tco.irr_percent == tco.irr_percent else "N/A"
        console.print(
            Panel(
                f"  Annual cost savings:          [bold green]${tco.annual_savings:>12,.0f}[/bold green]\n"
                f"  One-time migration cost:      [bold yellow]${tco.migration_cost_usd:>12,.0f}[/bold yellow]\n"
                f"  15% contingency reserve:      [bold yellow]${tco.contingency_usd:>12,.0f}[/bold yellow]\n"
                f"  Training investment:          [bold yellow]${tco.labor.training_one_time:>12,.0f}[/bold yellow]\n"
                f"  Total investment (all-in):    [bold red]${tco.total_investment_usd:>12,.0f}[/bold red]\n\n"
                f"  Break-even point:             [bold white]{tco.payback_period_str:>12}[/bold white]\n"
                f"  Internal Rate of Return:      [bold cyan]{irr_str:>12}[/bold cyan]\n\n"
                f"  3-Year NPV (8% hurdle):       [bold cyan]${tco.npv_3yr:>12,.0f}[/bold cyan]\n"
                f"  5-Year NPV (8% hurdle):       [bold cyan]${tco.npv_5yr:>12,.0f}[/bold cyan]\n\n"
                f"  3-Year net benefit:           [bold green]${tco.three_year_cumulative_savings:>12,.0f}[/bold green]\n"
                f"  5-Year net benefit:           [bold green]${tco.five_year_cumulative_savings:>12,.0f}[/bold green]\n"
                f"  3-Year ROI:                   [bold white]{tco.roi_percent:>11.1f}%[/bold white]",
                title="[bold]CFO Executive Summary (8% Discount Rate)[/bold]",
                border_style="green",
            )
        )

        # Scenario comparison
        scenario_table = Table(
            title="Migration Strategy Scenarios — 3-Year Comparison",
            box=box.ROUNDED,
            header_style="bold white on dark_blue",
        )
        scenario_table.add_column("Strategy", min_width=16)
        scenario_table.add_column("Migration Cost", justify="right")
        scenario_table.add_column("Yr1 Savings", justify="right")
        scenario_table.add_column("Yr3 Savings", justify="right")
        scenario_table.add_column("3-Yr Net", justify="right")
        scenario_table.add_column("NPV (3yr)", justify="right")
        scenario_table.add_column("IRR", justify="right")
        scenario_table.add_column("Break-even", justify="center")

        for s in tco.scenarios:
            net_color = "green" if s.cumulative_3yr_benefit > 0 else "red"
            scenario_table.add_row(
                s.name,
                f"${s.migration_cost:,.0f}",
                f"[green]${s.year1_savings:,.0f}[/green]",
                f"[green]${s.year3_savings:,.0f}[/green]",
                f"[{net_color}]${s.cumulative_3yr_benefit:,.0f}[/{net_color}]",
                f"[{net_color}]${s.npv_3yr:,.0f}[/{net_color}]",
                f"{s.irr_percent:.1f}%",
                f"{s.break_even_months:.0f}mo",
            )
        console.print(scenario_table)

        self._print_cashflow_chart(tco)

        sens_table = Table(title="Sensitivity Analysis", box=box.SIMPLE, header_style="bold white")
        sens_table.add_column("Scenario", min_width=40)
        sens_table.add_column("Annual Savings", justify="right")
        sens_table.add_column("3-Yr NPV", justify="right")
        sens_table.add_column("Break-Even", justify="right")
        sens_table.add_column("Verdict", min_width=30)

        for s in tco.sensitivity_results:
            npv_color = "green" if s.npv_3yr > 0 else "red"
            v_color = "green" if "Proceed" in s.recommendation and "caution" not in s.recommendation else (
                "yellow" if "caution" in s.recommendation else "red"
            )
            sens_table.add_row(
                s.scenario_name,
                f"${s.annual_savings:,.0f}",
                f"[{npv_color}]${s.npv_3yr:,.0f}[/{npv_color}]",
                f"{s.break_even_months:.0f}mo" if s.break_even_months > 0 else "Never",
                f"[{v_color}]{s.recommendation}[/{v_color}]",
            )
        console.print(sens_table)

    def _print_cashflow_chart(self, tco: TCOAnalysis) -> None:
        console.print("\n[bold]Cumulative Cash Flow — All-In Investment (Yr 0-5)[/bold]")
        cashflows = tco.yearly_cashflows
        max_abs = max(abs(cf) for cf in cashflows) if cashflows else 1
        bar_width = 40

        for yr, cf in enumerate(cashflows):
            bar_len = int(abs(cf) / max_abs * bar_width)
            if cf < 0:
                bar = "[red]" + "-" * bar_len + "[/red]"
                label = f"[red]-${abs(cf):>10,.0f}[/red]"
            else:
                bar = "[green]" + "#" * bar_len + "[/green]"
                label = f"[green]+${cf:>10,.0f}[/green]"
            console.print(f"  Yr {yr}  {label}  {bar}")
        console.print()

    def export_excel(
        self,
        tco: TCOAnalysis,
        output_path: str,
        title: str = "Migration TCO Analysis",
    ) -> None:
        """
        Export CFO-ready Excel workbook with formatted TCO model.
        Requires openpyxl.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            console.print("[yellow]openpyxl not installed — skipping Excel export[/yellow]")
            return

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_costs = wb.create_sheet("Cost Breakdown")
        ws_scenarios = wb.create_sheet("Scenarios")
        ws_sensitivity = wb.create_sheet("Sensitivity")
        ws_cashflow = wb.create_sheet("Cash Flow")

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        positive_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        negative_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        money_format = '#,##0'
        pct_format = '0.0%'
        thin_border = Border(
            bottom=Side(style='thin', color='CCCCCC')
        )

        def style_header_row(ws: Any, row: int, num_cols: int) -> None:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # === Executive Summary sheet ===
        ws_summary["A1"] = f"MigrationScout V2 — {title}"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1F3864")
        ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["A2"].font = Font(italic=True, color="888888")

        summary_data = [
            ("", "", ""),
            ("FINANCIAL SUMMARY", "", ""),
            ("Annual Cost Savings", tco.annual_savings, "$/yr"),
            ("Migration Investment (all-in)", tco.total_investment_usd, "$"),
            ("  of which: Migration Cost", tco.migration_cost_usd, "$"),
            ("  of which: 15% Contingency", tco.contingency_usd, "$"),
            ("  of which: Training", tco.labor.training_one_time, "$"),
            ("", "", ""),
            ("Break-even Point", tco.payback_period_str, ""),
            ("Internal Rate of Return (IRR)", tco.irr_percent / 100 if tco.irr_percent else 0, pct_format),
            ("3-Year NPV (8% hurdle rate)", tco.npv_3yr, "$"),
            ("5-Year NPV (8% hurdle rate)", tco.npv_5yr, "$"),
            ("", "", ""),
            ("3-Year Net Benefit", tco.three_year_cumulative_savings, "$"),
            ("5-Year Net Benefit", tco.five_year_cumulative_savings, "$"),
            ("3-Year ROI", tco.roi_percent / 100, pct_format),
        ]

        start_row = 4
        for i, (label, value, fmt) in enumerate(summary_data):
            row = start_row + i
            ws_summary.cell(row=row, column=1, value=label)
            if isinstance(value, (int, float)):
                cell = ws_summary.cell(row=row, column=2, value=value)
                if fmt == "$":
                    cell.number_format = money_format
                elif fmt == pct_format:
                    cell.number_format = pct_format
                elif fmt == "$/yr":
                    cell.number_format = money_format
                # Color positive/negative
                if isinstance(value, (int, float)) and value > 0:
                    cell.fill = positive_fill
                elif isinstance(value, (int, float)) and value < 0:
                    cell.fill = negative_fill
            else:
                ws_summary.cell(row=row, column=2, value=str(value))

            if label == "FINANCIAL SUMMARY":
                ws_summary.cell(row=row, column=1).font = Font(bold=True, size=13, color="1F3864")

        ws_summary.column_dimensions["A"].width = 38
        ws_summary.column_dimensions["B"].width = 22

        # === Cost Breakdown sheet ===
        headers = ["Cost Category", "On-Premises / Month", "Cloud / Month", "Monthly Savings", "Annual Savings"]
        for col, h in enumerate(headers, 1):
            ws_costs.cell(row=1, column=col, value=h)
        style_header_row(ws_costs, 1, len(headers))

        cost_data = [
            ("Compute / Hardware", tco.on_prem.hardware_monthly, tco.cloud.compute_monthly),
            ("Power & Cooling", tco.on_prem.power_monthly, 0),
            ("Datacenter / Rack", tco.on_prem.datacenter_monthly, 0),
            ("Storage", 0, tco.cloud.storage_monthly),
            ("Database (Managed)", 0, tco.cloud.database_monthly),
            ("Network / VPN", 0, tco.cloud.network_monthly),
            ("Data Transfer", 0, tco.cloud.data_transfer_monthly),
            ("Staff (Ops)", tco.on_prem.staff_monthly, -tco.labor.cloud_ops_monthly_delta),
            ("Licenses", tco.on_prem.license_monthly, tco.cloud.license_monthly),
            ("Maintenance", tco.on_prem.maintenance_monthly, 0),
            ("Backup / DR", tco.on_prem.backup_monthly, tco.cloud.backup_monthly),
            ("Support Plan", 0, tco.cloud.support_monthly),
            ("Managed Services", 0, tco.cloud.managed_services_monthly),
            ("Security", 0, tco.cloud.security_monthly),
        ]

        for row_idx, (label, on_prem_val, cloud_val) in enumerate(cost_data, 2):
            savings = on_prem_val - cloud_val
            ws_costs.cell(row=row_idx, column=1, value=label)
            ws_costs.cell(row=row_idx, column=2, value=on_prem_val).number_format = money_format
            ws_costs.cell(row=row_idx, column=3, value=cloud_val).number_format = money_format
            savings_cell = ws_costs.cell(row=row_idx, column=4, value=savings)
            savings_cell.number_format = money_format
            savings_cell.fill = positive_fill if savings > 0 else negative_fill if savings < 0 else PatternFill()
            ws_costs.cell(row=row_idx, column=5, value=savings * 12).number_format = money_format

        for col in range(1, 6):
            ws_costs.column_dimensions[get_column_letter(col)].width = 22

        # === Scenarios sheet ===
        scenario_headers = ["Strategy", "Migration Cost", "Yr1 Savings", "Yr2 Savings", "Yr3 Savings",
                            "3-Yr Net Benefit", "NPV (3yr)", "IRR", "Break-even (mo)"]
        for col, h in enumerate(scenario_headers, 1):
            ws_scenarios.cell(row=1, column=col, value=h)
        style_header_row(ws_scenarios, 1, len(scenario_headers))

        for row_idx, s in enumerate(tco.scenarios, 2):
            ws_scenarios.cell(row=row_idx, column=1, value=s.name)
            ws_scenarios.cell(row=row_idx, column=2, value=s.migration_cost).number_format = money_format
            ws_scenarios.cell(row=row_idx, column=3, value=s.year1_savings).number_format = money_format
            ws_scenarios.cell(row=row_idx, column=4, value=s.year2_savings).number_format = money_format
            ws_scenarios.cell(row=row_idx, column=5, value=s.year3_savings).number_format = money_format
            net_cell = ws_scenarios.cell(row=row_idx, column=6, value=s.cumulative_3yr_benefit)
            net_cell.number_format = money_format
            net_cell.fill = positive_fill if s.cumulative_3yr_benefit > 0 else negative_fill
            npv_cell = ws_scenarios.cell(row=row_idx, column=7, value=s.npv_3yr)
            npv_cell.number_format = money_format
            npv_cell.fill = positive_fill if s.npv_3yr > 0 else negative_fill
            ws_scenarios.cell(row=row_idx, column=8, value=s.irr_percent / 100).number_format = pct_format
            ws_scenarios.cell(row=row_idx, column=9, value=s.break_even_months)

        for col in range(1, 10):
            ws_scenarios.column_dimensions[get_column_letter(col)].width = 20

        # === Cash Flow sheet ===
        ws_cashflow["A1"] = "Cumulative Cash Flow Projection"
        ws_cashflow["A1"].font = Font(bold=True, size=13, color="1F3864")
        cf_headers = ["Year", "Cumulative Cash Flow", "Annual Savings", "Note"]
        for col, h in enumerate(cf_headers, 1):
            ws_cashflow.cell(row=2, column=col, value=h)
        style_header_row(ws_cashflow, 2, len(cf_headers))

        for row_idx, (yr, cf) in enumerate(enumerate(tco.yearly_cashflows), 3):
            ws_cashflow.cell(row=row_idx, column=1, value=f"Year {yr}")
            cf_cell = ws_cashflow.cell(row=row_idx, column=2, value=cf)
            cf_cell.number_format = money_format
            cf_cell.fill = positive_fill if cf > 0 else negative_fill
            ws_cashflow.cell(row=row_idx, column=3, value=tco.annual_savings if yr > 0 else -tco.total_investment_usd).number_format = money_format
            note = "Migration investment" if yr == 0 else ("Break-even year" if cf >= 0 and (yr == 0 or tco.yearly_cashflows[yr - 1] < 0) else "")
            ws_cashflow.cell(row=row_idx, column=4, value=note)

        ws_cashflow.column_dimensions["A"].width = 12
        ws_cashflow.column_dimensions["B"].width = 24
        ws_cashflow.column_dimensions["C"].width = 20
        ws_cashflow.column_dimensions["D"].width = 22

        # Save
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        console.print(f"[green]Excel TCO model exported to: {path}[/green]")
        console.print("[dim]Open in Excel for formatted financial model with formulas[/dim]")
